#!/usr/bin/env python3
"""Generate the sample Excel file for SM30 bulk import GUI tests."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

HEADERS = ["VKORG", "MATNR_V", "MATNR_B", "PFEP_RUN_TYPE"]
ROWS = [
    ["E001", "V2607010001", "B2607010002", "K"],
    ["E001", "V2607010003", "B2607010004", "A"],
    ["E001", "V2607010005", "B2607010006", "S1"],
    ["E001", "V2607010007", "B2607010008", "K"],
    ["E001", "V2607010009", "B2607010010", "A"],
    ["E001", "V2607010011", "B2607010012", "S1"],
    ["E001", "V2607010013", "B2607010014", "K"],
    ["E001", "V2607010015", "B2607010016", "A"],
    ["E001", "V2607010017", "B2607010018", "S1"],
    ["E001", "V2607010019", "B2607010020", "K"],
    ["E001", "V2607010021", "B2607010022", "A"],
    ["E001", "V2607010023", "B2607010024", "S1"],
    ["E001", "V2607010025", "B2607010026", "K"],
    ["E001", "V2607010027", "B2607010028", "A"],
    ["E001", "V2607010029", "B2607010030", "S1"],
    ["E001", "V2607010031", "B2607010032", "K"],
    ["E001", "V2607010033", "B2607010034", "A"],
    ["E001", "V2607010035", "B2607010036", "S1"],
    ["E001", "V2607010037", "B2607010038", "K"],
    ["E001", "V2607010039", "B2607010040", "A"],
]

OUTPUT = Path(__file__).resolve().parents[1] / "examples" / "data" / "pfepruntype_sample.xlsx"


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "PFEPRUNTYPE"

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    notes = wb.create_sheet("Notes")
    notes["A1"] = "Sample workbook for SM30 bulk import GUI"
    notes["A3"] = "Sheet PFEPRUNTYPE maps to view /WUE/PFEPRUNTYPE"
    notes["A4"] = "Column order: VKORG, MATNR_V, MATNR_B, PFEP_RUN_TYPE"

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
